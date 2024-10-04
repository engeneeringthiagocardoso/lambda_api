import json
import boto3
import os
import tempfile
import csv
import time
import pdfkit
import unicodedata
import re

from PIL import Image, ImageDraw, ImageFont, ImageSequence
from docx import Document
from pptx import Presentation
from botocore.exceptions import ClientError
from openpyxl import load_workbook
from fpdf import FPDF
from PyPDF2 import PdfReader, PdfWriter

s3 = boto3.client('s3')
transcribe = boto3.client('transcribe')
eventbridge = boto3.client('events')

def lambda_handler(event, context):
    try:
        print("Received event:")
        print("Received event:", json.dumps(event, indent=2))
        # Configuração do bucket S3
        s3_bucket = ""
        s3_bucket_source = ""

        # Detalhes do evento recebido
        detail = event['detail']
        path = os.path.join(detail['path'], os.path.basename(detail['path_source']))  # Concatena o path e o path_source
        print(f"path: {path}")
        path_source = detail['path_source']
        metadata = detail['metadata']
        file_type = path_source.split('.')[-1].lower()

        # Processar o arquivo baseado no tipo
        if file_type in ['mp3', 'mp4', 'wav']:
            process_audio_video(s3_bucket_source, s3_bucket, path_source, file_type, metadata, path)
        # elif file_type in ['jpg', 'jpeg', 'png', 'gif']:
        #     convert_document_to_image_and_metadata(s3_bucket_source, s3_bucket, path_source, file_type, metadata, path)
        # elif file_type == 'zip':
        #     process_scorm(s3_bucket_source, s3_bucket, path_source, path)
        elif file_type in ['pdf', 'doc', 'docx', 'ppt', 'pptx', 'xls', 'xlsx', 'html', 'csv', 'jpg', 'jpeg', 'png', 'gif','xsd']:
            print("Entrei no file_type:")
            process_and_split_pdf(s3_bucket_source, s3_bucket, path_source, file_type, metadata, path)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")

        return {
            'statusCode': 200,
            'body': json.dumps(f'File processed: {path_source}')
        }

    except Exception as e:
        print(f"Error processing file {path_source}: {str(e)}")
        return {
            'statusCode': 500,
            'body': json.dumps({'error': str(e)})
        }

def process_and_split_pdf(s3_bucket_source, s3_bucket, path_source, file_type, metadata, path):
    try:
        with tempfile.TemporaryDirectory() as tmpdirname:
            local_file_path = os.path.join(tmpdirname, os.path.basename(path_source))
            print("Vou baixar")
            s3.download_file(s3_bucket_source, path_source, local_file_path)
            print("Baixei")
            # Converter para PDF, se necessário
            if file_type not in ['pdf']:
                local_file_path = convert_to_pdf(local_file_path, file_type)

            # Verificar o tamanho do arquivo PDF e dividir se necessário
            if os.path.getsize(local_file_path) > 49 * 1024 * 1024:
                pdf_files = split_pdf(local_file_path)
            else:
                pdf_files = [local_file_path]

            # Fazer upload de todos os arquivos (ou partes) gerados
            for pdf_file in pdf_files:
                pdf_file_name = f"{os.path.splitext(path)[0]}_part{pdf_files.index(pdf_file) + 1}.pdf"
                s3.upload_file(pdf_file, s3_bucket, pdf_file_name)
                save_metadata(s3_bucket, pdf_file_name, metadata)
                print(f"Arquivo {pdf_file_name} enviado para S3")

            # Salvar o metadata.json no S3
            # save_metadata(s3_bucket, path, metadata)

    except ClientError as e:
        print(f"Failed to process document: {e}")

def convert_to_pdf(local_file_path, file_type):
    print("Entrei no convert_to_pdf:")
    print(f"local_file_path: {local_file_path}")
    print(f"file_type: {file_type}")

    pdf_file_path = f"{os.path.splitext(local_file_path)[0]}.pdf"
    
    if file_type in ['jpg', 'jpeg', 'png', 'gif']:
        image = Image.open(local_file_path)
        image.convert('RGB').save(pdf_file_path)
    elif file_type in ['doc', 'docx']:
        # Para .doc e .docx, você pode usar bibliotecas como python-docx ou docx2pdf.
        # Aqui é um exemplo usando python-docx:
        document = Document(local_file_path)
        document.save(pdf_file_path)  # Esta linha é apenas ilustrativa; python-docx não salva diretamente em PDF.
        # Para conversão real para PDF, você pode precisar de bibliotecas como `docx2pdf`.
    elif file_type in ['ppt', 'pptx']:
        presentation = Presentation(local_file_path)
        # Convertendo slides para imagens e depois para PDF
        images = []
        for slide in presentation.slides:
            img = Image.new('RGB', (1920, 1080), color='white')
            d = ImageDraw.Draw(img)
            d.text((10, 10), "Slide Content Here", fill='black')  # Substitua com o conteúdo real
            images.append(img)
        images[0].save(pdf_file_path, save_all=True, append_images=images[1:])
    elif file_type in ['xls', 'xlsx','xsd']:
        process_excel(local_file_path, pdf_file_path)
    elif file_type in ['csv','xsd']:
        process_csv(local_file_path, pdf_file_path)
    elif file_type in ['html']:
        print("Entrei no html:")
        convert_html_to_pdf(local_file_path, pdf_file_path)
    
    return pdf_file_path

def convert_html_to_pdf(html_path, pdf_file_path):
    # O caminho do binário `wkhtmltopdf` na camada Lambda
    path_wkhtmltopdf = '/opt/bin/wkhtmltopdf'  # `/opt` é o local das camadas no Lambda
    config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)

    # Carregar o conteúdo HTML
    with open(html_path, 'r', encoding='utf-8') as file:
        html_content = file.read()

    # Adicionar o CSS para controle de quebra de página no <head>
    page_break_css = """
    <style>
      .page-break {
          page-break-before: always;
      }
      @media print {
          body {
              width: 100%;
          }
      }
      @page {
          size: A4;
          margin: 1cm;
      }
      html, body {
          zoom: 0.75; /* Ajusta o zoom para que mais conteúdo caiba na página */
      }
      img {
          max-width: 100%; /* Garante que as imagens não ultrapassem a largura da página */
      }
    </style>
    """

    # Inserir o CSS no <head> do HTML (ou criar um <head> se não existir)
    if "</head>" in html_content:
        html_content = html_content.replace("</head>", page_break_css + "</head>")
    else:
        html_content = "<head>" + page_break_css + "</head>" + html_content

    # Exemplo: inserir quebra de página antes de cada <h1> (você pode ajustar isso conforme necessário)
    html_content = html_content.replace("<h1>", '<div class="page-break"></div><h1>')

    # Salvar o HTML modificado temporariamente
    modified_html_path = '/tmp/modified_html.html'
    with open(modified_html_path, 'w', encoding='utf-8') as modified_file:
        modified_file.write(html_content)

    options = {
        'page-size': 'A1',  # Define o tamanho da página como A4
        'encoding': 'UTF-8',
        'no-outline': None,  # Remove o contorno no PDF
        'print-media-type': '',  # Forçar o uso do CSS de mídia de impressão
        'zoom': '0.75',  # Reduz o zoom para caber mais conteúdo por página
    }

    # Converter o arquivo HTML modificado em PDF
    pdfkit.from_file(modified_html_path, pdf_file_path, configuration=config, options=options)

    print(f"HTML renderizado e convertido para PDF: {pdf_file_path}")
    
def split_pdf(pdf_file_path):
    pdf_reader = PdfReader(pdf_file_path)
    pdf_files = []
    pdf_writer = PdfWriter()
    
    for i in range(len(pdf_reader.pages)):
        pdf_writer.add_page(pdf_reader.pages[i])
        temp_pdf_path = f"{pdf_file_path}_part{i+1}.pdf"
        with open(temp_pdf_path, 'wb') as out_pdf:
            pdf_writer.write(out_pdf)
        if os.path.getsize(temp_pdf_path) > 30 * 1024 * 1024 or i == len(pdf_reader.pages) - 1:
            pdf_files.append(temp_pdf_path)
            pdf_writer = PdfWriter()  # Reset writer for next part
    
    return pdf_files

def save_metadata(s3_bucket, path, metadata):
    metadata_file_name = f"{os.path.splitext(path)[0]}.metadata.json"
    metadata_content = json.dumps(metadata)

    s3.put_object(
        Bucket=s3_bucket,
        Key=f"{metadata_file_name}",
        Body=metadata_content
    )
    print(f"Metadata {metadata_file_name} uploaded to S3")

def save_metadata_transcribe(s3_bucket, path, metadata):
    metadata_file_name = f"{os.path.splitext(path)[0]}.txt"
    metadata_content = json.dumps(metadata)

    s3.put_object(
        Bucket=s3_bucket,
        Key=f"{metadata_file_name}",
        Body=metadata_content
    )
    print(f"Metadata {metadata_file_name} uploaded to S3")


def upload_image_to_s3(tmpdirname, image_file_name, s3_bucket, path):
    print("Dentro do upload_image_to_s3")
    s3.upload_file(os.path.join(tmpdirname, image_file_name), s3_bucket, f"{path}{image_file_name}")
    print(f"Image {image_file_name} uploaded to S3")

def process_csv(file_path, pdf_file_path):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    with open(file_path, mode='r') as file:
        csv_reader = csv.reader(file)
        for row in csv_reader:
            line = ', '.join(row)
            pdf.cell(200, 10, txt=line, ln=True)
    
    pdf.output(pdf_file_path)

def process_excel(file_path, pdf_file_path):
    wb = load_workbook(filename=file_path)
    sheet = wb.active
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    for row in sheet.iter_rows(values_only=True):
        line = ', '.join([str(cell) for cell in row])
        pdf.cell(200, 10, txt=line, ln=True)

    pdf.output(pdf_file_path)

def sanitize_name(name):
    # Remove acentos e caracteres especiais
    sanitized = unicodedata.normalize('NFKD', name).encode('ascii', 'ignore').decode('ascii')
    # Substitui espaços por underscores e remove caracteres não permitidos
    sanitized = re.sub(r'[^a-zA-Z0-9._-]', '_', sanitized)
    return sanitized
def process_audio_video(s3_bucket_source, s3_bucket, path_source, file_type, metadata, path):
    try:
        with tempfile.TemporaryDirectory() as tmpdirname:
            local_file_path = os.path.join(tmpdirname, os.path.basename(path_source))
            
            # Baixar o arquivo de áudio ou vídeo do bucket source
            s3.download_file(s3_bucket_source, path_source, local_file_path)

            # Sanitizar os nomes
            transcribe_job_unique_name = f"transcription_{int(time.time() * 1000)}_{sanitize_name(os.path.basename(path_source).replace('.', '_'))}"
            transcribe_job_name = f"transcription_{sanitize_name(os.path.basename(path_source).replace('.', '_'))}"
            transcribe_output_key = f"transcriptions/{transcribe_job_name}.json"
            transcribe_output_url = f"s3://{s3_bucket}/{transcribe_output_key}"

            print(f"path: {path}")
            print(f"transcribe_output_key: {transcribe_output_key}")
            print(f"transcribe_output_url: {transcribe_output_url}")

            # Remover a extensão do caminho do arquivo, independentemente do formato
            original_path = os.path.splitext(path)[0]

            # Adicionar o caminho original sem a extensão ao metadado
            metadata['original_path'] = path
            save_metadata_transcribe(s3_bucket, transcribe_output_key, metadata)
            start_transcription_job(s3_bucket_source, path_source, transcribe_job_unique_name, s3_bucket, transcribe_output_key, file_type)

    except ClientError as e:
        print(f"Failed to process audio/video file: {e}")

def start_transcription_job(s3_bucket_source, path_source, transcribe_job_unique_name, s3_bucket, transcribe_output_key, file_type):
    try:
        response = transcribe.start_transcription_job(
            TranscriptionJobName=transcribe_job_unique_name,
            Media={'MediaFileUri': f's3://{s3_bucket_source}/{path_source}'},
            MediaFormat=file_type,  # Utiliza o file_type para definir o formato
            LanguageCode='pt-BR',  # Pode ajustar o código de linguagem conforme necessário
            OutputBucketName=s3_bucket,
            OutputKey=transcribe_output_key
        )
        print(f"Transcription job started: {transcribe_job_unique_name}")
    except ClientError as e:
        print(f"Failed to start transcription job: {e}")

def list_files_in_directory(directory):
    files_list = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            full_path = os.path.join(root, file)
            files_list.append(full_path)
            print("list_files_in_directory")
            print(full_path)  # Exibe o caminho completo do arquivo
    return files_list
